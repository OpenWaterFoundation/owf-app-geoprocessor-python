# pandas_util - utility functions related to the pandas library
# ________________________________________________________________NoticeStart_
# GeoProcessor
# Copyright (C) 2017-2020 Open Water Foundation
# 
# GeoProcessor is free software:  you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
# 
#     GeoProcessor is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
# 
#     You should have received a copy of the GNU General Public License
#     along with GeoProcessor.  If not, see <https://www.gnu.org/licenses/>.
# ________________________________________________________________NoticeEnd___

import os
import pandas as pd
from openpyxl import load_workbook
import xlwt

from geoprocessor.core.DataStore import DataStore
import geoprocessor.util.io_util as io_util


def create_data_frame_from_datastore_with_sql(sql_query: str, datastore_obj: DataStore):
    """
    Creates a pandas data frame object by running a SQL query on a DataFrame object.

    Args:
        sql_query (str):  a SQL query to get the correct table out of the DataStore object
        datastore_obj (obj): the DataStore object

    Returns:
        A pandas data frame object. This is the object that support the GeoProcessor's Table object.
    """

    df = pd.read_sql(sql_query, datastore_obj.engine)
    return df


def create_data_frame_from_delimited_file(file_path: str, delimiter: str):
    """
    Creates a pandas data frame object from a delimited file.

    Args:
        file_path (str): the full pathname to a delimited file
        delimiter (str): the delimiter symbol (often times is a comma)

    Returns:
        A pandas data frame object. This is the object that support the GeoProcessor's Table object.
    """

    # Define and return the data frame from the delimited file. Pass the delimited file's delimiter as the sep argument.
    df = pd.read_csv(file_path, sep=delimiter)
    return df


def create_data_frame_from_excel(excel_workbook_path: str, excel_worksheet_name: str = None):
    """
    Creates a pandas data frame object from an excel file.

    Args:
        excel_workbook_path (str): the full pathname to the excel file
        excel_worksheet_name (str): the name of the worksheet to read from the excel workbook

    Returns:
        A pandas data frame object. This is the object that support the GeoProcessor's Table object.
    """

    # If an excel worksheet is defined, define and return the data frame from the worksheet.
    if excel_worksheet_name:
        df = pd.read_excel(excel_workbook_path, sheet_name=excel_worksheet_name)
        return df

    # If an excel worksheet is not defined, define and return the data frame from the FIRST worksheet.
    else:
        df = pd.read_excel(excel_workbook_path)
        return df


def create_excel_workbook_obj(excel_workbook_path: str):
    """
    Creates a pandas excel workbook object from an excel file.

    Args:
        excel_workbook_path (str): the full pathname to the excel file

    Returns:
        A pandas excel workbook object. This object allows for the GeoProcessor to read the properties of the Excel
         workbook. For example, it can return the names of all of the available worksheets.
    """

    # Define and return the pandas excel workbook object from the excel file.
    xl = pd.ExcelFile(excel_workbook_path)
    return xl


def write_df_to_delimited_file(df, output_file_full_path: str, include_col_list: [str], include_index: bool,
                               delimiter: str = ",") -> None:
    """
    Writes a pandas data frame object to a delimited file.

    Args:
        df (object): the pandas data frame object to write
        output_file_full_path (str): the full pathname to an output delimited file
        include_col_list (list of strings): A list of Table columns to write to the delimited file
        include_index (bool): If TRUE, write the index column. If FALSE, exclude the index column.
        delimiter (str): the delimiter symbol to use in the output delimited file. Must be a single character. Default
         value is a comma.

    Returns: None
    """

    # Write the pandas data frame to a csv file.
    df.to_csv(output_file_full_path, columns=include_col_list, index=include_index, sep=delimiter)


def write_df_to_excel(df, excel_workbook_path: str, excel_worksheet_name: str, include_col_list: [str],
                      include_index: bool) -> None:
    """
    Writes a pandas data frame object to an excel file.

    Args:
        df (object): the pandas data frame object to write
        excel_workbook_path (str): the full pathname to an excel workbook (either existing or non-existing)
        excel_worksheet_name (str): the worksheet name to write to (either existing or non-existing)
        include_col_list (list of strings): A list of Table columns to write to the excel file
        include_index (bool): If TRUE, write out column names. If FALSE, do not write column names.

    Returns: None
    """

    # Removes the default styling of the table (provided in the pandas library). The package has been moved over time.
    try:
        import pandas.io.formats.excel
        pandas.io.formats.excel.header_style = None

    except Exception:
        pass

    try:
        import pandas.formats.format
        pandas.formats.format.header_style = None

    except Exception:
        pass

    try:
        import pandas.core.format
        pandas.core.format.header_style = None

    except Exception:
        pass

    # If the output excel file already exists, take into consideration the current file format and the current
    # worksheets.
    if os.path.exists(excel_workbook_path):

        # TODO egiles 2018-04-25 Currently this function does not work. Need to fix.
        # Write the table to an existing excel file in XLS format.
        if io_util.get_extension(excel_workbook_path).upper() == ".XLS":

            # Set the writer object.
            writer = pd.ExcelWriter(excel_workbook_path, engine = 'xlwt')

            book = xlwt.Workbook(excel_workbook_path)
            writer.book = book

            # Write the df to the excel workbook with the given worksheet name.
            df.to_excel(writer, sheet_name=excel_worksheet_name, index=include_index, columns=include_col_list)
            writer.save()

        # Write the table to an existing excel file in XLSX format.
        else:

            # REf: https://stackoverflow.com/questions/20219254/
            # how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas

            # Set the writer object.
            writer = pd.ExcelWriter(excel_workbook_path, engine="openpyxl")

            # If applicable, inform the writer object of the already-existing excel workbook.
            book = load_workbook(excel_workbook_path)
            writer.book = book

            # If applicable, inform the writer object of the already-existing excel worksheets.
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            # Write the df to the excel workbook with the given worksheet name.
            df.to_excel(writer, sheet_name=excel_worksheet_name, index=include_index, columns=include_col_list)
            writer.save()

    # If the output excel file does not already exists, configure which excel file version to use.
    else:

        # Set the writer object.
        writer = pd.ExcelWriter(excel_workbook_path)

        # Write the df to the excel workbook with the given worksheet name.
        df.to_excel(writer, sheet_name=excel_worksheet_name, index=include_index, columns=include_col_list)
        writer.save()
